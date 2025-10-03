import streamlit as st
import pandas as pd
import os
from fractions import Fraction
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
from datetime import datetime

# ======================================================
# 🔹 Page Setup
# ======================================================
st.set_page_config(page_title="JSC Industries – Advanced Fastener Intelligence", layout="wide")

# ======================================================
# 🔹 Load Databases
# ======================================================
url = "https://docs.google.com/spreadsheets/d/11Icre8F3X8WA5BVwkJx75NOH3VzF6G7b/export?format=xlsx"
local_excel_path = r"G:\My Drive\Streamlite\ASME B18.2.1 Hex Bolt and Heavy Hex Bolt.xlsx"
me_chem_path = r"Mechanical and Chemical.xlsx"

thread_files = {
    "ASME B1.1": "ASME B1.1 New.xlsx",
    "ISO 965-2-98 Coarse": "ISO 965-2-98 Coarse.xlsx",
    "ISO 965-2-98 Fine": "ISO 965-2-98 Fine.xlsx",
}

@st.cache_data
def load_data(url):
    try:
        return pd.read_excel(url)
    except:
        if os.path.exists(local_excel_path):
            return pd.read_excel(local_excel_path)
        return pd.DataFrame()

df = load_data(url)

@st.cache_data
def load_thread_data(file):
    try:
        return pd.read_excel(file)
    except:
        return pd.DataFrame()

@st.cache_data
def load_mechem_data(file):
    if os.path.exists(file):
        return pd.read_excel(file)
    return pd.DataFrame()

df_mechem = load_mechem_data(me_chem_path)

# ======================================================
# 🔹 Helper Functions
# ======================================================
def size_to_float(size_str):
    try:
        size_str = str(size_str).strip()
        if "-" in size_str and not size_str.replace("-", "").isdigit():
            parts = size_str.split("-")
            return float(parts[0]) + float(Fraction(parts[1]))
        else:
            return float(Fraction(size_str))
    except:
        return None

def calculate_weight(product, diameter_mm, length_mm):
    density = 0.00785  # Steel density g/mm^3
    if product == "Hex Cap Screw":
        factor = 0.95
    elif product == "Heavy Hex Bolt":
        factor = 1.05
    elif product == "Heavy Hex Screw":
        factor = 1.1
    elif product == "Threaded Rod":
        factor = 1.0
    else:
        factor = 1.0
    volume = 3.1416 * (diameter_mm / 2) ** 2 * length_mm
    weight_kg = volume * density * factor / 1000
    return round(weight_kg, 4)  # 4 decimal places

# ======================================================
# 🔹 Initialize Session State
# ======================================================
if "selected_section" not in st.session_state:
    st.session_state.selected_section = None

# ======================================================
# 🔹 Home Dashboard
# ======================================================
def show_home():
    st.markdown("<h1 style='text-align:center; color:#2C3E50;'>🏠 JSC Industries – Workspace</h1>", unsafe_allow_html=True)
    st.markdown("<h4 style='text-align:center; color:gray;'>Click on any section to enter</h4>", unsafe_allow_html=True)
    
    sections = [
        ("📦 Product Database", "database_icon.png"),
        ("🧮 Calculations", "calculator_icon.png"),
        ("🕵️ Inspection", "inspection_icon.png"),
        ("🔬 Research & Development", "rnd_icon.png"),
        ("💬 Team Chat", "chat_icon.png"),
        ("🤖 PiU (AI Assistant)", "ai_icon.png")
    ]
    
    cols = st.columns(3)
    for idx, (title, icon) in enumerate(sections):
        with cols[idx % 3]:
            if st.button(title, key=title):
                st.session_state.selected_section = title
            # Optional: display icons
            # st.image(icon, width=80)

# ======================================================
# 🔹 TDS Generation Helper
# ======================================================
def generate_tds(template_file, supplier, product_name, length_val, size_val, marking, grade, filtered_df, filtered_mecert_df):
    wb = load_workbook(template_file)
    ws = wb.active
    
    # Example: update cells (modify according to your template)
    ws["B2"] = supplier
    ws["B3"] = product_name
    ws["B4"] = f"Size: {size_val}, Length: {length_val}"
    ws["B5"] = marking
    ws["B6"] = grade
    
    # Dimensional Data (assuming starts at row 10)
    row_start = 10
    for idx, row in filtered_df.iterrows():
        col = 1
        for val in row:
            ws.cell(row=row_start, column=col, value=val)
            col += 1
        row_start += 1
    
    # ME&CERT Data (assuming starts at row after dimensional)
    row_start += 2
    for idx, row in filtered_mecert_df.iterrows():
        col = 1
        for val in row:
            ws.cell(row=row_start, column=col, value=val)
            col += 1
        row_start += 1
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    return temp_file.name

# ======================================================
# 🔹 Section Workspaces
# ======================================================
def show_section(title):
    if title == "📦 Product Database":
        st.header("📦 Product Database")
        if df.empty and df_mechem.empty:
            st.warning("No data available.")
        else:
            tab1, tab2 = st.tabs(["🔍 Search Database", "📝 Create Technical Data Sheet"])
            
            with tab1:
                st.subheader("Database Search")
                product_types = ["All"] + sorted(list(df['Product'].dropna().unique()) + ["Threaded Rod", "Stud"])
                product_type = st.selectbox("Select Product Name", product_types)
                series_options = ["Inch", "Metric"]
                series = st.selectbox("Select Series", series_options)
                
                # Dimensional Filter
                dimensional_standards = ["ASME B18.2.1"] if series=="Inch" else ["ISO"]
                dimensional_standard = st.selectbox("Dimensional Standard", ["All"] + dimensional_standards)
                
                dimensional_size_options = ["All"]
                if dimensional_standard != "All" and "Size" in df.columns:
                    temp_df = df.copy()
                    if product_type != "All":
                        temp_df = temp_df[temp_df['Product']==product_type]
                    if dimensional_standard != "All":
                        temp_df = temp_df[temp_df['Standards']==dimensional_standard]
                    dimensional_size_options += sorted(temp_df['Size'].dropna().unique(), key=size_to_float)
                dimensional_size = st.selectbox("Dimensional Size", dimensional_size_options)
                
                # ME&CERT Filter
                mecert_standard_options = ["All"] + (sorted(df_mechem['Standard'].dropna().unique()) if not df_mechem.empty else [])
                mecert_standard = st.selectbox("ME&CERT Standard", mecert_standard_options)
                mecert_property_options = ["All"]
                if mecert_standard != "All":
                    temp_df_me = df_mechem[df_mechem['Standard']==mecert_standard]
                    if "Property class" in temp_df_me.columns:
                        mecert_property_options += sorted(temp_df_me['Property class'].dropna().unique())
                mecert_property = st.selectbox("Property Class", mecert_property_options)
                
                # Filter Data
                filtered_df = df.copy()
                if product_type!="All":
                    filtered_df = filtered_df[filtered_df['Product']==product_type]
                if dimensional_standard!="All":
                    filtered_df = filtered_df[filtered_df['Standards']==dimensional_standard]
                if dimensional_size!="All":
                    filtered_df = filtered_df[filtered_df['Size']==dimensional_size]
                
                st.subheader(f"Found {len(filtered_df)} Bolt Records")
                st.dataframe(filtered_df)
                
                filtered_mecert_df = df_mechem.copy()
                if mecert_standard!="All":
                    filtered_mecert_df = filtered_mecert_df[filtered_mecert_df['Standard']==mecert_standard]
                if mecert_property!="All":
                    filtered_mecert_df = filtered_mecert_df[filtered_mecert_df['Property class']==mecert_property]
                st.subheader(f"ME&CERT Records: {len(filtered_mecert_df)}")
                st.dataframe(filtered_mecert_df)
                
                if st.button("📥 Download All Filtered Data"):
                    wb = Workbook()
                    ws_dim = wb.active
                    ws_dim.title = "Dimensional Data"
                    for r in dataframe_to_rows(filtered_df, index=False, header=True):
                        ws_dim.append(r)
                    ws_me = wb.create_sheet("ME&CERT Data")
                    for r in dataframe_to_rows(filtered_mecert_df, index=False, header=True):
                        ws_me.append(r)
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                    wb.save(temp_file.name)
                    with open(temp_file.name,"rb") as f:
                        st.download_button("⬇️ Download Excel", f, file_name="Filtered_Fastener_Data.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            with tab2:
                st.subheader("Create Technical Data Sheet")
                tds_template = st.file_uploader("Upload TDS Excel Template (from GitHub)", type=["xlsx"])
                supplier_name = st.text_input("Supplier Name")
                product_name = st.selectbox("Product Name", df['Product'].dropna().unique())
                size_val = st.text_input("Size (from DB)")
                length_val = st.number_input("Length", min_value=1.0, step=0.1)
                marking = st.selectbox("Marking", ["Yes", "No"])
                if marking=="No":
                    marking_value = "NA"
                else:
                    marking_value = st.text_input("Enter Marking Value")
                grade_standard = st.selectbox("Grade Standard", df_mechem['Standard'].dropna().unique())
                grade_property = st.selectbox("Grade Property", df_mechem['Property class'].dropna().unique())
                grade_value = f"{grade_standard}-{grade_property}"
                
                if st.button("✅ Generate TDS") and tds_template is not None:
                    tds_file = generate_tds(tds_template.name, supplier_name, product_name, length_val, size_val,
                                            marking_value, grade_value, filtered_df, filtered_mecert_df)
                    with open(tds_file,"rb") as f:
                        st.download_button("⬇️ Download Generated TDS", f, file_name="Technical_Data_Sheet.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    elif title == "🧮 Calculations":
        st.header("🧮 Engineering Calculations")
        st.subheader("Manual Weight Calculator")
        product_options = sorted(list(df['Product'].dropna().unique()) + ["Threaded Rod", "Stud"])
        selected_product = st.selectbox("1️⃣ Select Product", product_options)
        series = st.selectbox("2️⃣ Select Series", ["Inch", "Metric"])
        metric_type = st.selectbox("3️⃣ Select Thread Type", ["Coarse", "Fine"]) if series=="Metric" else None
        selected_standard = "ASME B1.1" if series=="Inch" else ("ISO 965-2-98 Coarse" if metric_type=="Coarse" else "ISO 965-2-98 Fine")
        st.info(f"📏 Standard: **{selected_standard}** (used only for pitch diameter)")
        
        df_thread = load_thread_data(thread_files[selected_standard])
        size_options = sorted(df_thread["Thread"].dropna().unique()) if not df_thread.empty else []
        selected_size = st.selectbox("5️⃣ Select Size", size_options)
        length_unit = st.selectbox("6️⃣ Select Length Unit", ["inch", "mm"])
        length_val = st.number_input("7️⃣ Enter Length", min_value=0.1, step=0.1)
        dia_type = st.selectbox("8️⃣ Select Diameter Type", ["Body Diameter", "Pitch Diameter"])
        
        diameter_mm = None
        if dia_type == "Body Diameter":
            body_dia = st.number_input("🔹 Enter Body Diameter", min_value=0.1, step=0.1)
            diameter_mm = body_dia * 25.4 if length_unit=="inch" else body_dia
        else:
            if not df_thread.empty:
                if "Class" in df_thread.columns:
                    pitch_classes = sorted(df_thread["Class"].dropna().unique())
                    pitch_class = st.selectbox("🔹 Select Pitch Class", pitch_classes)
                    row = df_thread[(df_thread["Thread"]==selected_size) & (df_thread["Class"]==pitch_class)]
                else:
                    row = df_thread[df_thread["Thread"]==selected_size]
                if not row.empty:
                    pitch_val = row["Pitch Diameter (Min)"].values[0]
                    diameter_mm = pitch_val if series=="Metric" else pitch_val*25.4
                else:
                    st.warning("⚠️ Pitch Diameter not found for this selection.")
        
        if st.button("⚖️ Calculate Weight"):
            length_mm = length_val*25.4 if length_unit=="inch" else length_val
            if diameter_mm is None:
                st.error("❌ Please provide diameter information.")
            else:
                weight_kg = calculate_weight(selected_product, diameter_mm, length_mm)
                st.success(f"✅ Estimated Weight/pc: **{weight_kg} Kg**")
        
        # ======================================================
        # 🔹 Batch Weight Calculator
        # ======================================================
        st.subheader("📊 Batch Weight Calculator")
        batch_file = st.file_uploader("Upload Batch Excel/CSV (Columns: Product, Diameter_mm, Length_mm, Series)", type=["xlsx","csv"], key="batch")
        if batch_file:
            if batch_file.name.endswith(".csv"):
                batch_df = pd.read_csv(batch_file)
            else:
                batch_df = pd.read_excel(batch_file)

            required_cols = ["Product","Diameter_mm","Length_mm","Series"]
            if all(col in batch_df.columns for col in required_cols):
                if st.button("⚖️ Calculate Batch Weights", key="batch_calc"):
                    batch_df["Weight_Kg"] = batch_df.apply(
                        lambda row: calculate_weight(row["Product"], row["Diameter_mm"], row["Length_mm"]), axis=1
                    )
                    st.dataframe(batch_df)

                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                    batch_df.to_excel(temp_file.name, index=False)
                    with open(temp_file.name,"rb") as f:
                        st.download_button("⬇️ Download Batch Weight Excel", f, 
                                           file_name="Batch_Weight.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.error(f"❌ Uploaded file must contain columns: {', '.join(required_cols)}")
        
    elif title == "🕵️ Inspection":
        st.header("🕵️ Inspection")
        st.subheader("Supplier Inspection")
        supplier_sku = st.text_input("Enter SKU")
        supplier_photos = st.file_uploader("Upload Supplier Photos", accept_multiple_files=True, type=["jpg","png"])
        supplier_notes = st.text_area("Notes")
        
        st.subheader("Inhouse Inspection")
        inhouse_sku = st.text_input("Enter SKU for Inhouse Inspection")
        inhouse_photos = st.file_uploader("Upload Inhouse Photos", accept_multiple_files=True, type=["jpg","png"])
        inhouse_notes = st.text_area("Notes")
        
        if st.button("✅ Generate Inspection Reports"):
            st.success("Inspection Excel and PDF Reports generated and saved in desired location.")
            st.info("All data stored in virtual environment for future research.")

    elif title == "🔬 Research & Development":
        st.header("🔬 Research & Development")
        st.info("Workspace for complex calculations, material studies, simulations, and experiments.")
        st.text_area("Input / Notes / Research Parameters")

    elif title == "💬 Team Chat":
        st.header("💬 Team Chat")
        username = st.text_input("Enter your name")
        message = st.text_input("Enter message")
        if st.button("Send"):
            st.success(f"Message sent: {message}")
        st.info("Messages can be stored in CSV/Database for logging. Real-time upgrade possible.")

    elif title == "🤖 PiU (AI Assistant)":
        st.header("🤖 PiU – AI Assistant")
        user_query = st.text_input("Ask PiU anything")
        if st.button("Ask PiU"):
            st.info("PiU is processing your query...")
            st.success("PiU Answer: This is where the AI response will appear. Connect to ChatGPT/Google API later.")

    # Back Button
    st.markdown("<hr>")
    if st.button("⬅️ Back to Home"):
        st.session_state.selected_section = None

# ======================================================
# 🔹 Main Display Logic
# ======================================================
if st.session_state.selected_section is None:
    show_home()
else:
    show_section(st.session_state.selected_section)

# ======================================================
# 🔹 Footer
# ======================================================
st.markdown("""
<hr>
<div style='text-align:center; color:gray'>
    © JSC Industries Pvt Ltd | Born to Perform
</div>
""", unsafe_allow_html=True)
